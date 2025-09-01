# 計算部分所需函數
import os
import sys

sys.path.append("./")
import datetime
import traceback
from copy import copy

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tabulate import tabulate

import app.services.utils as utils


def setting_txt():
    result = []
    with open("SETTING.txt", "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            try:
                line = int(line)
                result.append(line)
            except:
                result.append(line)
    setting = {"database_path": "", "pn_col": "", "fir_col": ""}
    for i, d in zip(setting, result[:3]):
        setting[i] = d
    return setting


data = setting_txt()
dict_setting = setting_txt()
state1 = [False, False, False]


class Setup_window(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


class SubWindow_controller(QtWidgets.QMainWindow):
    def __init__(self):
        # in python3, super(Class, self).xxx = super().xxx
        super().__init__()
        self.ui = Ui_SubWindow()
        self.ui.setupUi(self)
        self.setup_control()
        self.ui.path_text.setText(data["database_path"])
        self.ui.pn_text.setText(data["pn_col"])
        self.ui.firstcol_text.setText(str(data["fir_col"]))

    def setup_control(self):
        self.ui.path_check.clicked.connect(self.p_check)
        self.ui.path_tool.clicked.connect(self.open_file)
        self.ui.pn_check.clicked.connect(self.partnum_check)
        self.ui.firstcol_check.clicked.connect(self.fircol_check)
        self.ui.save_button.clicked.connect(self.save_setting)
        self.ui.close_button.clicked.connect(self.closeEvent)

    def p_check(self):
        self.ui.path_text.setText(data["database_path"])
        self.ui.path_tool.setEnabled(False)
        self.ui.path_text.setEnabled(False)
        if self.ui.path_check.isChecked():
            self.ui.path_tool.setEnabled(True)
            self.ui.path_text.setEnabled(True)
        print(self.ui.path_check.isChecked())

    def partnum_check(self):
        self.ui.pn_text.setText(data["pn_col"])
        self.ui.pn_text.setEnabled(False)
        if self.ui.pn_check.isChecked():
            self.ui.pn_text.setEnabled(True)

    def fircol_check(self):
        self.ui.firstcol_text.setText(str(data["fir_col"]))
        self.ui.firstcol_text.setEnabled(False)
        if self.ui.firstcol_check.isChecked():
            self.ui.firstcol_text.setEnabled(True)

    def open_file(self):
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "Open folder",
        )  # start path
        self.ui.path_text.setText(folder_path)

    # 點擊save時保存當前的設置內容，在點擊close時套用前一次的配置
    def state_keep(self):
        print(
            self.ui.path_check.isChecked(),
            self.ui.pn_check.isChecked(),
            self.ui.firstcol_check.isChecked(),
        )
        state1[0] = self.ui.path_check.isChecked()
        state1[1] = self.ui.pn_check.isChecked()
        state1[2] = self.ui.firstcol_check.isChecked()

    def save_setting(self):
        dict_setting["database_path"] = self.ui.path_text.toPlainText()
        dict_setting["pn_col"] = self.ui.pn_text.toPlainText().upper()
        dict_setting["fir_col"] = int(self.ui.firstcol_text.toPlainText())
        # 留存前一次的保存配置
        self.state_keep()
        self.close()

    def closeEvent(self, event):
        self.ui.path_text.setText(dict_setting["database_path"])
        self.ui.pn_text.setText(dict_setting["pn_col"])
        self.ui.firstcol_text.setText(str(dict_setting["fir_col"]))
        self.ui.path_check.setChecked(state1[0])
        self.ui.path_text.setEnabled(state1[0])
        self.ui.pn_check.setChecked(state1[1])
        self.ui.pn_text.setEnabled(state1[1])
        self.ui.firstcol_check.setChecked(state1[2])
        self.ui.firstcol_text.setEnabled(state1[2])
        self.close()

    # setwindow保存時同步更改subwindow內容，並且重置check爲False
    def set_data(self):
        self.ui.path_text.setText(data["database_path"])
        self.ui.pn_text.setText(data["pn_col"])
        self.ui.firstcol_text.setText(str(data["fir_col"]))
        self.ui.path_check.setChecked(False)
        self.ui.pn_check.setChecked(False)
        self.ui.firstcol_check.setChecked(False)


class MainWindow_controller(QtWidgets.QMainWindow):
    def __init__(self):
        # in python3, super(Class, self).xxx = super().xxx
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_control()
        self.preview_setting()

    def setup_control(self):
        self.ui.bompath_tool_1.clicked.connect(self.open_file_1)
        self.ui.review_button.clicked.connect(self.add_substitute_review)
        self.ui.y_radio.toggled.connect(self.chose)
        self.ui.r_radio.toggled.connect(self.chose)
        self.ui.n_radio.toggled.connect(self.chose)
        self.ui.custom_radio.toggled.connect(self.chose)

        self.ui.bompath_tool_2.clicked.connect(self.open_file_2)
        self.ui.import_button.clicked.connect(self.run_import)

        self.ui.update_button.clicked.connect(self.run_update)

    def preview_setting(self):
        if self.ui.custom_radio.isChecked():
            self.ui.database_path.setText(dict_setting["database_path"])
            self.ui.pn_text.setText(
                dict_setting["pn_col"] + str(dict_setting["fir_col"])
            )
        else:
            self.ui.database_path.setText(data["database_path"])
            self.ui.pn_text.setText(data["pn_col"] + str(data["fir_col"]))

    def open_file_1(self):
        filename, filetype = QFileDialog.getOpenFileName(
            self, "Open file"
        )  # start path

        self.ui.bompath_text_1.setText(filename)

    def open_file_2(self):
        filename, filetype = QFileDialog.getOpenFileName(
            self, "Open file"
        )  # start path

        self.ui.bompath_text_2.setText(filename)

    def chose(self):
        self.ui.review_button.clicked.disconnect()

        if self.ui.y_radio.isChecked():
            self.ui.review_button.clicked.connect(self.add_substitute_review)
        elif self.ui.r_radio.isChecked():
            self.ui.review_button.clicked.connect(self.result_bom_review)
        elif self.ui.n_radio.isChecked():
            self.ui.review_button.clicked.connect(self.sys_bom_review)
        elif self.ui.custom_radio.isChecked():
            self.ui.review_button.clicked.connect(self.custom_bom_review)
            self.ui.custom_button.setEnabled(True)
            self.preview_setting()
            return
        self.ui.custom_button.setEnabled(False)
        self.preview_setting()

    def add_substitute_review(self):
        print("tiptop")
        try:
            if utils.check_file(data["database_path"]) == None:
                pass
            else:
                self.ui.display_1.append(utils.check_file(data["database_path"]))
                return
            wb_mapping, ws_mapping = utils.load(
                "%s/mapping.xlsx" % (data["database_path"])
            )
            dict_mapping = utils.to_dict(ws_mapping)
            p = self.ui.bompath_text_1.toPlainText()
            p = self.format_change(p)

            wb_bom, ws_bom = utils.load(p)
            self.ui.display_1.append("開始BOM REVIEW\n")
            self.ui.display_1.append(p.split("/")[-1])
            l = chr(ws_bom.max_column + 65)
            ws_bom["%s%d" % (l, data["fir_col"] - 1)].value = "CE Comment"
            for cell in ws_bom[l]:
                fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                cell.fill = fill
            others = []
            previous = ""
            print(ws_bom[data["pn_col"]][data["fir_col"] - 1].value)
            for action, num, comment in zip(
                ws_bom["A"][data["fir_col"] - 1 :],
                ws_bom[data["pn_col"]][data["fir_col"] - 1 :],
                ws_bom[l][data["fir_col"] - 1 :],
            ):
                if type(num.value) != str:
                    continue
                elif len(num.value) != 16:
                    continue
                else:
                    if action.value == "Add Substitute":
                        if num.value in dict_mapping:
                            if (
                                dict_mapping[num.value][2].value == previous
                                and dict_mapping[num.value][2].value.upper() != "AGREE"
                            ):
                                comment.value = "同上"
                                comment.fill = copy(dict_mapping[num.value][2].fill)
                            else:
                                comment.value = dict_mapping[num.value][2].value
                                comment.fill = copy(dict_mapping[num.value][2].fill)
                                comment.font = copy(dict_mapping[num.value][2].font)
                                previous = dict_mapping[num.value][2].value
                        else:
                            others.append(num.value)

                    else:
                        if num.value in dict_mapping:
                            comment.value = dict_mapping[num.value][2].value
                            comment.fill = copy(dict_mapping[num.value][2].fill)
                            comment.font = copy(dict_mapping[num.value][2].font)
                            previous = dict_mapping[num.value][2].value
                        else:
                            others.append(num.value)

            p_list = p.split("/")
            p_list[-1] = "(%s)" % (datetime.date.today()) + p_list[-1]
            np = "/".join(p_list)
            wb_bom.save(np)
            os.remove(p)
            self.ui.display_1.append("\n共 %d 筆空白資料：\n" % len(others))
            for i in others:
                self.ui.display_1.append(i)
            self.ui.display_1.append(
                "\n填入完成\n-------------------------------------------------------------------------------\n\n"
            )
        except Exception as e:
            self.ui.display_1.append(
                "Error:\n"
                + str(traceback.format_exc())
                + "\n-------------------------------------------------------------------------------"
            )

    def sys_bom_review(self):
        print("xitong")
        try:
            if utils.check_file(data["database_path"]) == None:
                pass
            else:
                self.ui.display_1.append(utils.check_file(data["database_path"]))
                return
            wb_mapping, ws_mapping = utils.load(
                "%s/mapping.xlsx" % (data["database_path"])
            )
            dict_mapping = utils.to_dict(ws_mapping)
            p = self.ui.bompath_text_1.toPlainText()
            p = self.format_change(p)

            wb_bom, ws_bom = utils.load(p)
            self.ui.display_1.append("開始BOM REVIEW\n")
            self.ui.display_1.append(p.split("/")[-1])
            l = chr(ws_bom.max_column + 65)
            ws_bom["%s%d" % (l, data["fir_col"] - 1)].value = "CE Comment"
            for cell in ws_bom[l]:
                fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                cell.fill = fill
            others = []
            previous = ""
            print(ws_bom[data["pn_col"]][data["fir_col"] - 1].value)
            for action, num, comment in zip(
                ws_bom["A"][2:], ws_bom["C"][2:], ws_bom[l][2:]
            ):
                if type(num.value) != str:
                    continue
                elif len(num.value) != 16:
                    continue
                else:
                    if action.value == None:
                        if num.value in dict_mapping:
                            if (
                                dict_mapping[num.value][2].value == previous
                                and dict_mapping[num.value][2].value.upper() != "AGREE"
                            ):
                                comment.value = "同上"
                                comment.fill = copy(dict_mapping[num.value][2].fill)
                            else:
                                comment.value = dict_mapping[num.value][2].value
                                comment.fill = copy(dict_mapping[num.value][2].fill)
                                comment.font = copy(dict_mapping[num.value][2].font)
                                previous = dict_mapping[num.value][2].value
                        else:
                            others.append(num.value)

                    else:
                        if num.value in dict_mapping:
                            comment.value = dict_mapping[num.value][2].value
                            comment.fill = copy(dict_mapping[num.value][2].fill)
                            comment.font = copy(dict_mapping[num.value][2].font)
                            previous = dict_mapping[num.value][2].value

                        else:
                            others.append(num.value)

            p = p.split("/")
            p[-1] = "(%s)" % (datetime.date.today()) + p[-1]
            np = "/".join(p)
            wb_bom.save(np)
            self.ui.display_1.append("\n共 %d 筆空白資料：\n" % len(others))
            for i in others:
                self.ui.display_1.append(i)
            self.ui.display_1.append(
                "\n填入完成\n-------------------------------------------------------------------------------\n\n"
            )
        except Exception as e:
            self.ui.display_1.append(
                "Error:\n"
                + str(traceback.format_exc())
                + "\n-------------------------------------------------------------------------------"
            )

    def result_bom_review(self):
        print("result")
        try:
            if utils.check_file(data["database_path"]) == None:
                pass
            else:
                self.ui.display_1.append(utils.check_file(data["database_path"]))
                return
            wb_mapping, ws_mapping = utils.load(
                "%s/mapping.xlsx" % (data["database_path"])
            )
            dict_mapping = utils.to_dict(ws_mapping)
            p = self.ui.bompath_text_1.toPlainText()
            p = self.format_change(p)

            wb_bom, ws_bom = utils.load(p)
            self.ui.display_1.append("開始BOM REVIEW\n")
            self.ui.display_1.append(p.split("/")[-1])
            l = chr(ws_bom.max_column + 65)
            ws_bom["%s%d" % (l, 1)].value = "CE Comment"
            for cell in ws_bom[l]:
                fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                cell.fill = fill
            others = []
            for num, comment in zip(ws_bom["B"], ws_bom[l]):
                if type(num.value) != str:
                    continue
                elif len(num.value) != 16:
                    continue
                elif num.value in dict_mapping:
                    comment.value = dict_mapping[num.value][2].value
                    comment.fill = copy(dict_mapping[num.value][2].fill)
                else:
                    others.append(num.value)
            p_list = p.split("/")
            p_list[-1] = "(%s)" % (datetime.date.today()) + p_list[-1]
            np = "/".join(p_list)
            wb_bom.save(np)
            os.remove(p)
            self.ui.display_1.append("\n共 %d 筆空白資料：\n" % len(others))
            for i in others:
                self.ui.display_1.append(i)
            self.ui.display_1.append(
                "\n填入完成\n-------------------------------------------------------------------------------\n\n"
            )
        except Exception as e:
            self.ui.display_1.append(
                "Error:\n"
                + str(traceback.format_exc())
                + "\n-------------------------------------------------------------------------------"
            )

    def custom_bom_review(self):
        print("custom")
        try:
            if utils.check_file(dict_setting["database_path"]) == None:
                pass
            else:
                self.ui.display_1.append(
                    utils.check_file(dict_setting["database_path"])
                )
                return
            wb_mapping, ws_mapping = utils.load(
                "%s/mapping.xlsx" % (dict_setting["database_path"])
            )
            dict_mapping = utils.to_dict(ws_mapping)
            p = self.ui.bompath_text_1.toPlainText()

            wb_bom, ws_bom = utils.load(p)
            self.ui.display_1.append("開始BOM REVIEW\n")
            self.ui.display_1.append(p.split("/")[-1])
            l = chr(ws_bom.max_column + 65)
            ws_bom["%s%d" % (l, dict_setting["fir_col"] - 1)].value = "CE Comment"
            for cell in ws_bom[l]:
                fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                cell.fill = fill
            others = []
            for num, comment in zip(
                ws_bom[dict_setting["pn_col"]][dict_setting["fir_col"] - 1 :],
                ws_bom[l][dict_setting["fir_col"] - 1 :],
            ):
                print(num.value)
                if type(num.value) != str:
                    continue
                elif len(num.value) != 16:
                    continue
                elif num.value in dict_mapping:
                    comment.value = dict_mapping[num.value][2].value
                    comment.fill = copy(dict_mapping[num.value][2].fill)
                else:
                    others.append(num.value)
            p = p.split("/")
            p[-1] = "(%s)" % (datetime.date.today()) + p[-1]
            np = "/".join(p)
            wb_bom.save(np)
            self.ui.display_1.append("\n共 %d 筆空白資料：\n" % len(others))
            for i in others:
                self.ui.display_1.append(i)
            self.ui.display_1.append(
                "\n填入完成\n-------------------------------------------------------------------------------\n\n"
            )
        except Exception as e:
            self.ui.display_1.append(
                "Error:\n"
                + str(traceback.format_exc())
                + "\n-------------------------------------------------------------------------------"
            )

    def run_import(self):
        try:
            if utils.check_file(data["database_path"]) == None:
                pass
            else:
                self.ui.display_2.append(utils.check_file(data["database_path"]))
                return
            wb_mapping, ws_mapping = utils.load(
                "%s/mapping.xlsx" % (data["database_path"])
            )
            wb_mapping.save("mapping import-backup.xlsx")

            path = self.ui.bompath_text_2.toPlainText()
            wb_bom, ws_bom = utils.load(path)
            self.ui.display_2.append("Import to Database\n")
            wb_compare = Workbook()
            ws_dup = wb_compare.active
            ws_dup.append(["料號", "BOM", "Database"])

            dict_mapping = utils.to_dict(ws_mapping)
            count = 0
            dup_count = 0
            dict_add = {}
            previous = ""
            l = chr(ws_bom.max_column + 64)
            for a, b, c, d in zip(
                ws_bom[data["pn_col"]][data["fir_col"] - 1 :],
                ws_bom[chr(ord(data["pn_col"]) + 1)][data["fir_col"] - 1 :],
                ws_bom[chr(ord(data["pn_col"]) + 2)][data["fir_col"] - 1 :],
                ws_bom[l][data["fir_col"] - 1 :],
            ):
                if d.value is None:
                    continue
                elif a.value is None:
                    continue
                elif len(a.value) != 16:
                    continue

                if a.value.strip() in dict_mapping:
                    if d.value != dict_mapping[a.value][2].value:
                        if d.value == "同上":
                            if dict_mapping[a.value][2].value != previous:
                                ws_dup.append(
                                    [a.value, previous, dict_mapping[a.value][2].value]
                                )
                                dup_count += 1
                                continue
                            else:
                                continue
                        else:
                            ws_dup.append(
                                [a.value, d.value, dict_mapping[a.value][2].value]
                            )
                            dup_count += 1
                            previous = d.value
                            continue
                else:
                    if d.value == "同上":
                        d.value = previous
                    ws_mapping.append([a.value, b.value, c.value, d.value])
                    ws_mapping["D"][-1].fill = copy(d.fill)
                    ws_mapping["D"][-1].font = copy(d.font)
                    dict_add[a.value] = (b, c, d)
                    count += 1
                previous = d.value

            self.ui.display_2.append("新增 %d 笔,compare警示%d" % (count, dup_count))
            ws_mapping.protection.enable()
            wb_mapping.save("%s/mapping.xlsx" % (data["database_path"]))
            wb_mapping.save("%smapping.xlsx" % (datetime.date.today()))
            if dup_count > 0:
                wb_compare.save(
                    "compare(%s).xlsx" % (path.split("/")[-1].split(".")[0])
                )

            wb_maintain, ws_maintain = utils.load(
                "%s/maintain.xlsx" % (data["database_path"])
            )
            wb_maintain.save("maintain-backup.xlsx")

            c = utils.to_maintain(wb_maintain, dict_add)
            wb_maintain.save("%s/maintain.xlsx" % (data["database_path"]))
            wb_maintain.save("%smaintain.xlsx" % (datetime.date.today()))
            self.ui.display_2.append("\n工作表新增資料如下：\n")
            for i in c:
                self.ui.display_2.append("[%s] 新增 %d 筆資料" % (i, c[i]))
            self.ui.display_2.append(
                "\n-------------------------------------------------------------------------------"
            )

        except Exception as e:
            self.ui.display_2.append(
                "Error:\n"
                + str(traceback.format_exc())
                + "\n-------------------------------------------------------------------------------"
            )

    def run_update(self):
        try:
            if utils.check_file(data["database_path"]) == None:
                pass
            else:
                self.ui.display_3.append(utils.check_file(data["database_path"]))
                return
            wb_mapping, ws_mapping = utils.load(
                "%s/mapping.xlsx" % (data["database_path"])
            )
            wb_mapping.save("mapping-update-backup.xlsx")
            self.ui.display_3.append("UPDATE\n")
            for name in ["機構料件", "電子料(1)", "電子料(2)", "電子料(R,C)", "Others"]:
                upd = []
                new = []
                color = []

                print(name)
                wb_maintain, ws_maintain = utils.load(
                    "%s/maintain.xlsx" % (data["database_path"]), name
                )
                dict_maintain = utils.to_dict(ws_maintain)
                dict_mapping = utils.to_dict(ws_mapping)

                for i in dict_maintain:
                    if i in dict_mapping:
                        if dict_mapping[i][2].value != dict_maintain[i][2].value:
                            upd.append(
                                [i, dict_mapping[i][2].value, dict_maintain[i][2].value]
                            )
                            dict_mapping[i][2].value = dict_maintain[i][2].value

                        if dict_mapping[i][2].fill.start_color.rgb != dict_maintain[i][
                            2
                        ].fill.start_color.rgb or dict_mapping[i][2].font != copy(
                            dict_maintain[i][2].font
                        ):
                            dict_mapping[i][2].fill = copy(dict_maintain[i][2].fill)
                            dict_mapping[i][2].font = copy(dict_maintain[i][2].font)
                            color.append([i, dict_maintain[i][2].value])
                    else:
                        ws_mapping.append(
                            [
                                i,
                                dict_maintain[i][0].value,
                                dict_maintain[i][1].value,
                                dict_maintain[i][2].value,
                            ]
                        )
                        new.append([i, dict_maintain[i][2].value])
                # ws_mapping.protection.enable()
                # wb_mapping.save('\\gctfile.gigacomputing.intra\NR2B\NR2B6\共用資料區\BOM DATABASE/mapping.xlsx')
                # wb_mapping.save('%smapping.xlsx'%(datetime.date.today()))
                # wb_maintain.save('%smaintain.xlsx'%(datetime.date.today()))

                self.ui.display_3.append("【%s】：\n" % name)
                if len(new) != 0:
                    self.ui.display_3.append("新增 %s 筆資料：\n" % len(new))
                    # df = pd.DataFrame(new,columns = ['PartNum','Comment'])
                    # df.index=df.index+1
                    self.ui.display_3.append(
                        tabulate(new, headers=["PartNum", "Comment"], tablefmt="html")
                    )
                    # self.ui.display_3.append(tabulate(new, headers="keys", tablefmt = 'grid'))

                if len(upd) != 0:
                    self.ui.display_3.append("\n更新comment %s 筆資料：\n" % len(upd))
                    # df = pd.DataFrame(upd,columns = ['PartNum','Old_Comment','New_Comment'])
                    # df.index=df.index+1
                    self.ui.display_3.append(
                        tabulate(
                            upd,
                            headers=["PartNum", "Old_Comment", "New_Comment"],
                            tablefmt="html",
                        )
                    )

                if len(color) != 0:
                    self.ui.display_3.append(
                        "\n更新High Light %s 筆資料\n" % len(color)
                    )
                    # df = pd.DataFrame(color,columns = ['PartNum','Comment'])
                    # df.index=df.index+1
                    self.ui.display_3.append(
                        tabulate(color, headers=["PartNum", "Comment"], tablefmt="html")
                    )

            ws_mapping.protection.enable()
            wb_mapping.save("%s/mapping.xlsx" % (data["database_path"]))
            wb_mapping.save("%smapping.xlsx" % (datetime.date.today()))
            wb_maintain.save("%smaintain.xlsx" % (datetime.date.today()))
            self.ui.display_3.append(
                "\n更新完成\n-------------------------------------------------------------------------------"
            )
        except Exception as e:
            self.ui.display_3.append(
                "Error:\n"
                + str(traceback.format_exc())
                + "\n-------------------------------------------------------------------------------"
            )

    def format_change(self, path):
        if path.split(".")[-1] == "xls":
            txt_path = path[:-3] + "txt"
            os.rename(path, txt_path)
            wb = Workbook()
            ws = wb.active
            with open(txt_path, "r") as f:
                for line in f:
                    l = (
                        repr(line)
                        .replace("\\n", "")
                        .replace("'", "")
                        .replace('"', "")
                        .split("\\t")
                    )
                    r = []
                    for i in l:
                        try:
                            i = int(i)
                            r.append(i)
                        except:
                            r.append(i)
                            continue
                    ws.append(r)
                    # print(repr(line))
                    # l.append([repr(line)])
            new_path = os.path.join(path, path[:-3] + "xlsx")
            wb.save(new_path)
            os.remove(txt_path)
        elif path.split(".")[-1] != "xlsx":
            self.ui.display_1.append("檔案格式錯誤")
        else:
            new_path = path

        return new_path
