import datetime
import os
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Protection
from openpyxl.utils import column_index_from_string


def check_database(database_path):
    """
    檢查 mapping.xlsx 和 maintain.xlsx 是否存在及是否被打開
    """
    mapping_file = os.path.join(database_path, "mapping.xlsx")
    maintain_file = os.path.join(database_path, "maintain.xlsx")
    lock_mapping_file = os.path.join(database_path, "~$mapping.xlsx")
    lock_maintain_file = os.path.join(database_path, "~$maintain.xlsx")
    # 檢查 mapping.xlsx 和 maintain.xlsx 是否存在
    if not os.path.exists(mapping_file):
        return "找不到 mapping.xlsx，請確認資料庫路徑設定正確！"
    if not os.path.exists(maintain_file):
        return "找不到 maintain.xlsx，請確認資料庫路徑設定正確！"
    # 檢查 maintain.xlsx 和 maintain.xlsx 是否被打開
    if os.path.exists(lock_maintain_file):
        return "請關閉 maintain.xlsx 後再執行！"
    if os.path.exists(lock_mapping_file):
        return "請關閉 mapping.xlsx 後再執行！"
    return None


def check_bom(bom_path):
    """
    確認 bom_path 不為空且檔案存在且格式正確
    """
    # 先檢查 bom_path
    if not bom_path:
        return "Error: \n\tBOM 檔案路徑為空！"
    bom_file = Path(bom_path)
    if not bom_file.exists():
        return f"Error: \n\tBOM 檔案不存在 \n\t\t→ {bom_path}"
    if bom_file.suffix.lower() not in [".xls", ".xlsx"]:
        return f"Error: \n\tBOM 檔案格式錯誤，請選擇 Excel 檔案 \n\t\t→ {bom_path}"


def read_files(bom_path, database_path):
    """
    讀取 BOM 原始 xls 或 xlsx 檔案並轉換為 DataFrame；同時加載 mapping.xlsx 為Series
    """
    suffix = Path(bom_path).suffix.lower()
    bom_df = pd.DataFrame()
    if suffix == ".xls":
        with open(bom_path, encoding="big5") as f:
            raw_data = [line.rstrip("\n") for line in f]
        bom_df = pd.DataFrame([line.split("\t") for line in raw_data])
    elif suffix == ".xlsx":
        bom_df = pd.read_excel(bom_path, header=None)

    mapping_df = pd.read_excel(f"{database_path}/mapping.xlsx")
    mapping_comment = mapping_df.set_index("料號")["說明"]
    return bom_df, mapping_comment


def load(path, name=None):
    """載入工作簿與工作表"""
    wb = load_workbook(path)
    ws = wb[name] if name else wb.active
    assert ws is not None  # 這裡告訴 Pylance ws 一定不是 None
    return wb, ws


# --- 下方為 comment 的處理邏輯 ---
def correct_comment_main(row):
    if row["Action"] == "Add":
        return row["raw_comment"]
    elif pd.notna(row["main_comment"]) and row["raw_comment"] == row["main_comment"]:
        return "同主料"
    return row["raw_comment"]


def correct_comment_system(row):
    if pd.notna(row["主件料號"]):
        return row["raw_comment"]
    elif pd.notna(row["main_comment"]) and row["raw_comment"] == row["main_comment"]:
        return "同主料"
    return row["raw_comment"]


COMMENT_METHODS = {
    "main": correct_comment_main,
    "system": correct_comment_system,
}


def correct_comment(row, method) -> Any:
    return COMMENT_METHODS[method](row)


# ??? comment 處理邏輯 ???P


def columns_from_string(col_name: str) -> int:
    """將 Excel 欄位名稱轉換為 0-based 的欄位索引"""
    return column_index_from_string(col_name) - 1


def hightlight_comment(mapping_path, new_path, col, row):
    # 重新讀取比對完成的 excel
    wb_mapping, ws_mapping = load(mapping_path)
    wb_bom, ws_bom = load(new_path)

    # 設定 comment 欄位為黃色
    last_col = ws_bom.max_column
    last_row = ws_bom.max_row
    for cells in ws_bom.iter_cols(min_col=last_col, max_col=last_col, max_row=last_row):
        for cell in cells:
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # 將 mapping 中非黃底的料號取出
    warning_dict = {}
    for r in ws_mapping.iter_rows(min_row=2):
        if r[-1].fill.start_color.rgb != "FFFFFF00":
            warning_dict[r[0].value] = copy(r[-1].fill)

    # 將 BOM 中的料號與 mapping 中的非黃底料號比對，若有符合則套用填色
    for r in ws_bom.iter_rows(min_row=row, min_col=col + 1):
        if r[0].value in warning_dict:
            r[-1].fill = warning_dict[r[0].value]

    # 儲存修改後的 BOM 檔案
    wb_bom.save(new_path)


def path_detail(bom_path):
    """
    獲取檔案名稱與父目錄
    """
    p = Path(bom_path)
    parent_dir = p.parent
    filename = p.name
    filename_stem = p.stem
    return parent_dir, filename, filename_stem


def check_and_log(msg, api):
    """
    檢查 msg 是否有值，若有則輸出並回傳 True
    """
    if msg:
        api.logs("review", msg)
        return True
    return False


def other_logs(api, e=None, new_filename=None):
    if new_filename:
        api.logs("review", f"Review completed!\nSaved as 【{new_filename}】")
    elif e:
        api.logs(
            "review",
            f"Review failed: \n\t檔案規格不符：\n\t\t缺少必要欄位或辨識值 -> 【{e}】",
        )
    else:
        api.logs("review", "Error: \n\t此檔案可能非 Result BOM，請重新確認檔案規格 ！")
    api.logs("review", "\n----------------------------------------\n")


def save_to_excel(df, parent_dir, filename_stem):
    """
    儲存 DataFrame 為新的 Excel 檔案
    檔名格式： (YYYY-MM-DD){原檔名}.xlsx
    """
    today = datetime.date.today().strftime("%Y-%m-%d")
    new_filename = f"({today}){filename_stem}.xlsx"
    new_path = parent_dir / new_filename
    df.to_excel(new_path, index=False, header=False)
    return new_path, new_filename


def find_unmatched(df, mapping_comment, col_idx, api):
    col_name = df.columns[col_idx]
    unmatched = df.loc[
        (~df.iloc[:, col_idx].isin(mapping_comment.index))
        & (df.iloc[:, col_idx].map(lambda x: isinstance(x, str) and len(x) == 16)),
        col_name,
    ]
    if not unmatched.empty:
        api.logs("review", f"\t共 {len(unmatched)} 筆待維護料號:")
        api.logs("review", "\n".join(f"\t\t{pn}" for pn in unmatched))


# 更具料號進行maintain各工作表的匹配
def match(pn):
    electronic_1 = [
        "10DK",
        "10DP",
        "10DS",
        "10DW",
        "10DZ",
        "10GL",
        "10HP",
        "10IF",
        "10IT",
        "10LT",
        "10TA",
        "11IF",
        "11IT",
        "11TA",
        "11TC",
        "11TS",
        "11TT",
        "11WC",
        "11WP",
        "11WR",
        "10TC",
        "10DE",
        "10TT",
    ]
    electronic_2 = [
        "10CP",
        "10CT",
        "10DL",
        "10DR",
        "10FF",
        "10FP",
        "10LB",
        "10LC",
        "10LF",
        "10LI",
        "10LN",
        "10OC",
        "10OD",
        "10XH",
        "10XT",
        "11BL",
        "11DL",
        "11DR",
        "11FP",
        "11LC",
        "11LF",
        "11XC",
        "11XF",
        "11XR",
    ]
    electronic_3 = [
        "10RC",
        "10RH",
        "10RN",
        "10RS",
        "10CE",
        "10CG",
        "10CL",
        "10CM",
        "10CN",
        "10CO",
        "10UA",
        "10WW",
        "10WR",
        "10WP",
        "10WA",
        "11BB",
        "11CE",
        "11CL",
        "11CO",
        "11RH",
    ]
    mechanism_list = [
        "10AC",
        "10NH",
        "10NR",
        "10SA",
        "10SB",
        "10SC",
        "10SL",
        "10SM",
        "10SR",
        "10ST",
        "11AC",
        "11NH",
        "11NR",
        "11SA",
        "11SC",
        "11SI",
        "11SM",
        "11SR",
        "12AC",
        "12AI",
        "12KR",
        "10KS",
    ]

    maintain_dict = {
        "電子料(1)": electronic_1,
        "電子料(2)": electronic_2,
        "電子料(R,C)": electronic_3,
        "機構料件": mechanism_list,
    }
    for name in maintain_dict:
        if pn[:4] in maintain_dict[name]:
            return name
    return "Others"


# 將對應料號相關資料寫入maintain各個資料表中
def to_maintain(wb, d):
    count = {
        "電子料(1)": 0,
        "電子料(2)": 0,
        "電子料(R,C)": 0,
        "機構料件": 0,
        "Others": 0,
    }
    for i in d:
        name = match(i)
        ws_maintain = wb[name]
        if i not in [p.value for p in ws_maintain["A"]]:
            ws_maintain.append(
                [i, d[i][0].value, d[i][1].value, d[i][2].value, datetime.date.today()]
            )
            ws_maintain["A"][-1].protection = Protection(locked=False)
            ws_maintain["D"][-1].protection = Protection(locked=False)
            ws_maintain["D"][-1].fill = copy(d[i][2].fill)
            ws_maintain["D"][-1].font = copy(d[i][2].font)
            ws_maintain["E"][-1].protection = Protection(locked=False)
            ws_maintain.protection.enable()
            count[name] += 1
    return count


# 轉換資料為字典形態(鍵為料號[A列],值為儲存格位置[C列])
def to_dict(ws):
    pn = [pn.value.strip() for pn in ws["A"]][1:]
    d1 = ws["B"][1:]
    d2 = ws["C"][1:]
    ce = ws["D"][1:]
    dict_i = {}
    for p, d1, d2, ce in zip(pn, d1, d2, ce):
        dict_i[p] = (d1, d2, ce)
    return dict_i


# 更新mapping時，選擇要上傳更新的maintain工作表
# def sheet_name():
#     sheet_name = input('''請輸入要更新的工作表對應代碼：
#                         1 : 機構料件
#                         2 : Jack
#                         3 : Laney
#                         4 : Andy
#                         5 : Others
# 請輸入代碼→''')
#     if sheet_name == '1':
#         return '機構料件'
#     elif sheet_name == '2':
#         return 'Jack'
#     elif sheet_name == '3':
#         return 'Laney'
#     elif sheet_name == '4':
#         return 'Andy'
#     elif sheet_name == '5':
#         return 'Others'
#     else:
#         return '輸入錯誤'

# # 確認開始執行程式
# def start():
#     while True:
#         answer = input('請輸入 start 開始執行 \n')
#         if answer != 'start':
#             os._exit(0)
#         else:
#             break

# # 確認退出程式
# def exit():
#     while True:
#         answer = input('退出請按 Enter \n')
#         if answer == '':
#             os._exit(0)
