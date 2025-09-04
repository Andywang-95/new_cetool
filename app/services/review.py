# %%
import datetime
from copy import copy
from pathlib import Path

import pandas as pd

# import utils
from openpyxl.styles import PatternFill

import app.services.utils as utils

# %%


def main_review_service(data, bom_path, api):
    msg = utils.check_bom(bom_path)
    if msg:
        api.logs("review", msg)
        return
    # 獲取檔案名稱與父目錄
    p = Path(bom_path)
    parent_dir = p.parent
    filename = p.name
    filename_stem = p.stem

    # 檢查資料庫檔案狀態
    msg = utils.check_database(data["database_path"])
    if msg:
        api.logs("review", msg)
        return

    api.logs("review", f"Starting review for 【{filename}】...")
    try:
        # 讀出BOM檔案
        bom_df = utils.change_df(bom_path)
        # 複製前 5 行原始資料
        header_rows = bom_df.iloc[:6].copy()
        header_rows.loc[header_rows.index[-1], header_rows.shape[1]] = "CE Comment"
        # 設定第 6 行是header，後續是資料內容，並新增 comment 欄位，
        bom_data = bom_df.iloc[6:].copy()
        bom_data.columns = bom_df.iloc[5]
        bom_data["group"] = (bom_data["Action"] == "Add").cumsum()
        # 讀出mapping檔案
        mapping_df = pd.read_excel(f"{data['database_path']}/mapping.xlsx")
        mapping_comment = mapping_df.set_index("料號")["說明"]
        # 計算不匹配料號
        unmatched = bom_data.loc[~bom_data["Number"].isin(mapping_df["料號"]), "Number"]
        print(unmatched)
        if unmatched.any():
            api.logs("review", "\t待維護料號:")
        for part_num in unmatched:
            if len(part_num) == 16:
                api.logs("review", f"\t\t{part_num}")
        # 分配對應料號的原始comment
        bom_data["raw_comment"] = bom_data["Number"].map(mapping_comment)
        # 取得群組主料的comment
        main_comment = bom_data[bom_data["Action"] == "Add"].set_index("group")[
            "raw_comment"
        ]
        # 分配主料comment到所有替料
        bom_data["main_comment"] = bom_data["group"].map(main_comment)
        # 根據條件填入 CE Comment
        bom_data["CE Comment"] = bom_data.apply(
            utils.correct_comment,
            axis=1,
            method="main",
        )
        bom_data.drop(columns=["group", "raw_comment", "main_comment"], inplace=True)
        final_df = pd.concat(
            [header_rows, pd.DataFrame(bom_data.values)], ignore_index=True
        )
        # 保存成新的 excel
        today = datetime.date.today().strftime("%Y-%m-%d")
        new_filename = f"({today}){filename_stem}.xlsx"
        new_path = parent_dir / new_filename
        final_df.to_excel(
            new_path,
            index=False,
            header=False,
        )
    except KeyError as e:
        api.logs(
            "review",
            f"Review failed: 缺少必要欄位或辨識值 -> 【{e}】\n請確認檔案是否符合規格 ！",
        )
        api.logs("review", "\n----------------------------------------\n")
        return
    # 重新讀取比對完成的 excel
    wb_mapping, ws_mapping = utils.load(f"{data['database_path']}/mapping.xlsx")
    wb_bom, ws_bom = utils.load(new_path)

    # 設定 comment 欄位為黃色
    last_col = ws_bom.max_column
    last_row = ws_bom.max_row
    for cells in ws_bom.iter_cols(min_col=last_col, max_col=last_col, max_row=last_row):
        for cell in cells:
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # 將 mapping 中非黃底的料號取出
    warning_dict = {}
    for row in ws_mapping.iter_rows(min_row=2):
        if row[-1].fill.start_color.rgb != "FFFFFF00":
            warning_dict[row[0].value] = copy(row[-1].fill)

    # 將 BOM 中的料號與 mapping 中的非黃底料號比對，若有符合則套用填色
    for row in ws_bom.iter_rows(min_row=7, min_col=3):
        if row[0].value in warning_dict:
            row[-1].fill = warning_dict[row[0].value]

    # 儲存修改後的 BOM 檔案
    wb_bom.save(new_path)
    api.logs("review", f"Review completed!\nSaved as 【{new_filename}】")
    api.logs("review", "\n----------------------------------------\n")


def result_review_service(data, bom_path, api):
    msg = utils.check_bom(bom_path)
    if msg:
        api.logs("review", msg)
        return
    # 獲取檔案名稱與父目錄
    p = Path(bom_path)
    parent_dir = p.parent
    filename = p.name
    filename_stem = p.stem

    # 檢查資料庫檔案狀態
    msg = utils.check_database(data["database_path"])
    if msg:
        api.logs("review", msg)
        return

    api.logs("review", f"Starting review for 【{filename}】...")
    # 讀出BOM檔案
    bom_data = utils.change_df(bom_path)
    if "Total count:" not in str(bom_data.iloc[0, 0]):
        api.logs("review", "Error: \n\t此檔案可能非 Result BOM，請重新確認檔案規格 ！")
        api.logs("review", "\n----------------------------------------\n")
        return

    # 新增說明欄位
    new_col_idx = bom_data.shape[1]
    bom_data[new_col_idx] = ""
    # 讀出mapping檔案
    mapping_df = pd.read_excel(f"{data['database_path']}/mapping.xlsx")
    mapping_comment = mapping_df.set_index("料號")["說明"]
    # 根據第二欄料號填入對應值
    bom_data[new_col_idx] = bom_data.iloc[:, 1].map(mapping_comment).fillna("")
    bom_data.iloc[2, new_col_idx] = "CE Comment"
    # 計算不匹配料號
    unmatched = bom_data.loc[~bom_data[1].isin(mapping_df["料號"]), 1]
    if unmatched.any():
        api.logs("review", "\t待維護料號:")
    for part_num in unmatched:
        if type(part_num) is str and len(part_num) == 16:
            api.logs("review", f"\t\t{part_num}")
    # 保存成新的 excel
    today = datetime.date.today().strftime("%Y-%m-%d")
    new_filename = f"({today}){filename_stem}.xlsx"
    new_path = parent_dir / new_filename
    bom_data.to_excel(
        new_path,
        index=False,
        header=False,
    )

    # 重新讀取比對完成的 excel
    wb_mapping, ws_mapping = utils.load(f"{data['database_path']}/mapping.xlsx")
    wb_bom, ws_bom = utils.load(new_path)

    # 設定 comment 欄位為黃色
    last_col = ws_bom.max_column
    last_row = ws_bom.max_row
    for cells in ws_bom.iter_cols(min_col=last_col, max_col=last_col, max_row=last_row):
        for cell in cells:
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # 將 mapping 中非黃底的料號取出
    warning_dict = {}
    for row in ws_mapping.iter_rows(min_row=2):
        if row[-1].fill.start_color.rgb != "FFFFFF00":
            warning_dict[row[0].value] = copy(row[-1].fill)

    # 將 BOM 中的料號與 mapping 中的非黃底料號比對，若有符合則套用填色
    for row in ws_bom.iter_rows(min_row=7, min_col=3):
        if row[0].value in warning_dict:
            row[-1].fill = warning_dict[row[0].value]

    # 儲存修改後的 BOM 檔案
    wb_bom.save(new_path)
    api.logs("review", f"Review completed!\nSaved as 【{new_filename}】")
    api.logs("review", "\n----------------------------------------\n")


def system_bom_review_service(data, bom_path, api):
    msg = utils.check_bom(bom_path)
    if msg:
        api.logs("review", msg)
        return
    # 獲取檔案名稱與父目錄
    p = Path(bom_path)
    parent_dir = p.parent
    filename = p.name
    filename_stem = p.stem

    # 檢查資料庫檔案狀態
    msg = utils.check_database(data["database_path"])
    if msg:
        api.logs("review", msg)
        return

    api.logs("review", f"Starting review for 【{filename}】...")
    try:
        # 讀出BOM檔案
        bom_data = utils.change_df(bom_path)
        # 設定第 6 行是header，後續是資料內容，並新增 comment 欄位，
        bom_data.columns = bom_data.iloc[0]
        bom_data["group"] = bom_data["主件料號"].notna().cumsum()
        # 讀出mapping檔案
        mapping_df = pd.read_excel(f"{data['database_path']}/mapping.xlsx")
        mapping_comment = mapping_df.set_index("料號")["說明"]
        # 計算不匹配料號
        unmatched = bom_data.loc[
            ~bom_data["元件/替代料號"].isin(mapping_df["料號"]), "元件/替代料號"
        ]
        print(unmatched)
        if unmatched.any():
            api.logs("review", "\t待維護料號:")
        for part_num in unmatched:
            if len(part_num) == 16:
                api.logs("review", f"\t\t{part_num}")
        # 分配對應料號的原始comment
        bom_data["raw_comment"] = bom_data["元件/替代料號"].map(mapping_comment)
        # 取得群組主料的comment
        main_comment = bom_data[bom_data["主件料號"].notna()].set_index("group")[
            "raw_comment"
        ]
        # 分配主料comment到所有替料
        bom_data["main_comment"] = bom_data["group"].map(main_comment)
        # 根據條件填入 CE Comment
        bom_data["CE Comment"] = bom_data.apply(
            utils.correct_comment, axis=1, method="system"
        )
        bom_data.drop(columns=["group", "raw_comment", "main_comment"], inplace=True)
        bom_data.iloc[0, bom_data.shape[1] - 1] = "CE Comment"
        # 保存成新的 excel
        today = datetime.date.today().strftime("%Y-%m-%d")
        new_filename = f"({today}){filename_stem}.xlsx"
        new_path = parent_dir / new_filename
        bom_data.to_excel(new_path, index=False, header=False)
    except KeyError as e:
        api.logs(
            "review",
            f"Review failed: 缺少必要欄位或辨識值 -> 【{e}】，請確認檔案是否符合規格 ！",
        )
        api.logs("review", "\n----------------------------------------\n")
        return
    # 重新讀取比對完成的 excel
    wb_mapping, ws_mapping = utils.load(f"{data['database_path']}/mapping.xlsx")
    wb_bom, ws_bom = utils.load(new_path)

    # 設定 comment 欄位為黃色
    last_col = ws_bom.max_column
    last_row = ws_bom.max_row
    for cells in ws_bom.iter_cols(min_col=last_col, max_col=last_col, max_row=last_row):
        for cell in cells:
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # 將 mapping 中非黃底的料號取出
    warning_dict = {}
    for row in ws_mapping.iter_rows():
        if row[-1].fill.start_color.rgb != "FFFFFF00":
            warning_dict[row[0].value] = copy(row[-1].fill)

    # 將 BOM 中的料號與 mapping 中的非黃底料號比對，若有符合則套用填色
    for row in ws_bom.iter_rows(min_row=7, min_col=3):
        if row[0].value in warning_dict:
            row[-1].fill = warning_dict[row[0].value]

    # 儲存修改後的 BOM 檔案
    wb_bom.save(new_path)
    api.logs("review", f"Review completed!\nSaved as 【{new_filename}】")
    api.logs("review", "\n----------------------------------------\n")


def custom_review_service(data, bom_path, api):
    pass


# %%
